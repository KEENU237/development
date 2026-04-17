"""
P&L Report Generator
Produces a richly formatted Excel workbook from the SQLite trade log.
"""

import logging
import os
from datetime import datetime

from config.settings import DATA_DIR

logger = logging.getLogger(__name__)


class PnLReportGenerator:
    """Generates Excel P&L reports using openpyxl."""

    def __init__(self, trade_log):
        self.trade_log = trade_log

    def generate(self, output_path: str = None) -> str:
        """
        Build and save the report.  Returns the file path, or '' on failure.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed — cannot generate report")
            return ""

        if not output_path:
            ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(DATA_DIR, f"pnl_{ts}.xlsx")

        os.makedirs(DATA_DIR, exist_ok=True)

        wb = openpyxl.Workbook()

        self._sheet_summary(wb, openpyxl)
        self._sheet_trade_log(wb, openpyxl)
        self._sheet_open_positions(wb, openpyxl)

        try:
            wb.save(output_path)
            logger.info(f"P&L report saved: {output_path}")
            return output_path
        except Exception as exc:
            logger.error(f"Report save failed: {exc}")
            return ""

    # ── Sheets ────────────────────────────────────────────────────────────────

    def _sheet_summary(self, wb, opx) -> None:
        ws  = wb.active
        ws.title = "Summary"
        s   = self.trade_log.get_daily_summary()

        H_FONT  = opx.styles.Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        H_FILL  = opx.styles.PatternFill("solid", fgColor="1F4E79")
        B_FONT  = opx.styles.Font(bold=True)
        G_FILL  = opx.styles.PatternFill("solid", fgColor="C6EFCE")
        R_FILL  = opx.styles.PatternFill("solid", fgColor="FFC7CE")
        Y_FILL  = opx.styles.PatternFill("solid", fgColor="FFEB9C")

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 22

        # Title
        ws["A1"] = "NSE F&O — Daily P&L Summary"
        ws["A1"].font = H_FONT
        ws["A1"].fill = H_FILL
        ws.merge_cells("A1:B1")

        rows = [
            ("Report Date",     s.get("date", "")),
            ("Total Trades",    s.get("total_trades", 0)),
            ("Open Trades",     s.get("open_trades", 0)),
            ("Closed Trades",   s.get("closed_trades", 0)),
            ("Win Trades",      s.get("win_trades", 0)),
            ("Loss Trades",     s.get("loss_trades", 0)),
            ("Win Rate %",      f"{s.get('win_rate', 0):.1f}%"),
            ("Gross P&L (₹)",   f"₹{s.get('gross_pnl', 0):,.2f}"),
        ]
        for i, (label, value) in enumerate(rows, start=2):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].font = B_FONT
            if label == "Gross P&L (₹)":
                fill = G_FILL if s.get("gross_pnl", 0) >= 0 else R_FILL
                ws[f"B{i}"].fill = fill
            if label == "Win Rate %":
                wr = s.get("win_rate", 0)
                ws[f"B{i}"].fill = G_FILL if wr >= 50 else R_FILL

    def _sheet_trade_log(self, wb, opx) -> None:
        ws = wb.create_sheet("Trade Log")

        H_FONT = opx.styles.Font(bold=True, color="FFFFFF")
        H_FILL = opx.styles.PatternFill("solid", fgColor="2E75B6")
        G_FILL = opx.styles.PatternFill("solid", fgColor="C6EFCE")
        R_FILL = opx.styles.PatternFill("solid", fgColor="FFC7CE")

        headers    = ["ID", "Symbol", "Strategy", "Expiry",
                      "Entry Time", "Exit Time",
                      "Entry ₹", "Exit ₹", "P&L ₹", "Status"]
        col_widths = [10, 12, 22, 12, 20, 20, 12, 12, 12, 10]

        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = H_FONT
            cell.fill      = H_FILL
            ws.column_dimensions[cell.column_letter].width = w

        for row_n, t in enumerate(self.trade_log.get_all_trades(), start=2):
            pnl     = t.get("realized_pnl") or 0
            status  = t.get("status", "")
            values  = [
                t.get("trade_id"),
                t.get("symbol"),
                t.get("strategy_name"),
                t.get("expiry"),
                (t.get("entry_time") or "")[:16],
                (t.get("exit_time")  or "")[:16],
                round(t.get("entry_premium") or 0, 2),
                round(t.get("exit_premium")  or 0, 2) if status == "CLOSED" else "",
                round(pnl, 2) if status == "CLOSED" else "",
                status,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_n, column=col, value=val)
            # Colour P&L cell
            pnl_cell = ws.cell(row=row_n, column=9)
            if status == "CLOSED":
                pnl_cell.fill = G_FILL if pnl >= 0 else R_FILL

    def _sheet_open_positions(self, wb, opx) -> None:
        ws = wb.create_sheet("Open Positions")

        H_FONT = opx.styles.Font(bold=True, color="FFFFFF")
        H_FILL = opx.styles.PatternFill("solid", fgColor="375623")

        headers    = ["ID", "Symbol", "Strategy", "Expiry",
                      "Entry Time", "Entry ₹", "Legs"]
        col_widths = [10, 12, 22, 12, 20, 12, 60]

        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = H_FONT
            cell.fill      = H_FILL
            ws.column_dimensions[cell.column_letter].width = w

        for row_n, t in enumerate(self.trade_log.get_open_trades(), start=2):
            legs_summary = "  |  ".join(
                f"{l['action']} {l['opt_type']} {int(l['strike'])} x{l['qty']} @ ₹{l['ltp']}"
                for l in t.get("legs", [])
            )
            values = [
                t.get("trade_id"),
                t.get("symbol"),
                t.get("strategy_name"),
                t.get("expiry"),
                (t.get("entry_time") or "")[:16],
                round(t.get("entry_premium") or 0, 2),
                legs_summary,
            ]
            for col, val in enumerate(values, start=1):
                ws.cell(row=row_n, column=col, value=val)
